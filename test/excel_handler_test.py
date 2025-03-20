import pytest
from unittest.mock import patch, MagicMock
import io
import openpyxl
from datetime import datetime
from fastapi import HTTPException
from core.handlers.excel_handler import generate_contacts_excel, generate_all_users_excel

# Mock data
mock_owner_email = "owner@example.com"
mock_linked_contacts = [
    {
        "user_id": "id1",
        "name": "John Doe",
        "email": "john@example.com",
        "phone": 1234567890,
        "relationship_type": "Friend"
    },
    {
        "user_id": "id2",
        "name": "Jane Doe",
        "email": "jane@example.com",
        "phone": 9876543210,
        "relationship_type": "Colleague"
    }
]

mock_contacts = [
    {
        "user_id": "id1",
        "name": "John Doe",
        "email": "john@example.com",
        "phone": 1234567890
    },
    {
        "user_id": "id2",
        "name": "Jane Doe",
        "email": "jane@example.com",
        "phone": 9876543210
    }
]

mock_relationships = [
    ("john@example.com", "id2", "Friend"),
    ("jane@example.com", "id1", "Colleague")
]


@pytest.fixture
def mock_get_linked_contacts():
    with patch('core.handlers.excel_handler.get_linked_contacts') as mock:
        yield mock


@pytest.fixture
def mock_contacts_collection():
    with patch('core.handlers.excel_handler.contacts_collection') as mock:
        yield mock


@pytest.fixture
def mock_pg_connection():
    with patch('core.handlers.excel_handler.get_pg_connection') as mock:
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_conn.__enter__.return_value = mock_conn
        mock_conn.cursor.return_value.__enter__.return_value = mock_cursor
        mock.return_value = mock_conn
        yield mock, mock_conn, mock_cursor


def test_generate_contacts_excel_success(mock_get_linked_contacts):
    # Mock get_linked_contacts to return test data
    mock_get_linked_contacts.return_value = mock_linked_contacts

    # Generate Excel file
    excel_data = generate_contacts_excel(mock_owner_email)

    # Check if xlsx data was returned
    assert isinstance(excel_data, bytes)
    assert len(excel_data) > 0

    # Parse the Excel file to check its content
    file_like_object = io.BytesIO(excel_data)
    workbook = openpyxl.load_workbook(file_like_object)
    worksheet = workbook["Linked Contacts"]

    # Check headers (row 4 - note: openpyxl uses 1-based indexing)
    assert worksheet.cell(row=4, column=1).value == "User ID"
    assert worksheet.cell(row=4, column=2).value == "Name"
    assert worksheet.cell(row=4, column=3).value == "Email"
    assert worksheet.cell(row=4, column=4).value == "Phone"
    assert worksheet.cell(row=4, column=5).value == "Relationship Type"

    # Check data rows
    assert worksheet.cell(row=5, column=1).value == "id1"
    assert worksheet.cell(row=5, column=2).value == "John Doe"
    assert worksheet.cell(row=5, column=5).value == "Friend"

    assert worksheet.cell(row=6, column=1).value == "id2"
    assert worksheet.cell(row=6, column=2).value == "Jane Doe"
    assert worksheet.cell(row=6, column=5).value == "Colleague"


def test_generate_contacts_excel_empty(mock_get_linked_contacts):
    # Mock get_linked_contacts to return empty list
    mock_get_linked_contacts.return_value = []

    # Check that it raises HTTPException
    with pytest.raises(HTTPException) as excinfo:
        generate_contacts_excel(mock_owner_email)

    assert excinfo.value.status_code == 404
    assert "No linked contacts found" in excinfo.value.detail


def test_generate_all_users_excel_success(mock_contacts_collection, mock_pg_connection):
    _, _, mock_cursor = mock_pg_connection

    # Mock MongoDB contacts_collection.find to return test data
    mock_contacts_cursor = MagicMock()
    mock_contacts_cursor.__iter__.return_value = mock_contacts
    mock_contacts_collection.find.return_value = mock_contacts_cursor

    # Mock PostgreSQL cursor to return test data for relationship counts
    mock_cursor.fetchall.side_effect = [
        [("john@example.com", 2), ("jane@example.com", 1)],  # First fetchall for counts
        mock_relationships  # Second fetchall for relationship details
    ]

    # Generate Excel file
    excel_data = generate_all_users_excel()

    # Check if xlsx data was returned
    assert isinstance(excel_data, bytes)
    assert len(excel_data) > 0

    # Parse the Excel file to check its content
    file_like_object = io.BytesIO(excel_data)
    workbook = openpyxl.load_workbook(file_like_object)

    # Check summary sheet
    summary_sheet = workbook["Users Summary"]
    assert summary_sheet.cell(row=4, column=1).value == "User ID"
    assert summary_sheet.cell(row=4, column=5).value == "Linked Contacts"
    assert "id1" in [summary_sheet.cell(row=5, column=1).value, summary_sheet.cell(row=6, column=1).value]

    # Check details sheet
    details_sheet = workbook["Contact Details"]
    assert details_sheet.cell(row=4, column=1).value == "Owner Email"
    assert details_sheet.cell(row=4, column=6).value == "Relationship Type"